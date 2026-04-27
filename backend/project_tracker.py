"""
Project Tracker — domain API.

Hierarchy of routes:
    /batches                                   GET, POST
    /batches/{id}/semesters                    GET, POST
    /semesters/{id}/courses                    GET, POST
    /courses/{id}/coordinators                 GET, POST
    /courses/{id}/projects                     GET, POST
    /projects/{id}                             GET, PUT, DELETE
    /projects/{id}/upload-excel                POST  (one row → one project's team)
    /students/{id}                             PUT, DELETE
    /download-template                         GET   (per-project XLSX layout)
"""

from __future__ import annotations

import os
from io import BytesIO
from typing import Any, Optional

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel
from supabase import Client, create_client

from jwt_auth import (
    verify_token,
    require_coordinator_for_course,
    require_coordinator_for_project,
)

# ── Supabase client ───────────────────────────────────────────────────────────
_URL = os.getenv("SUPABASE_URL", "")
_KEY = os.getenv("SUPABASE_SERVICE_KEY", "")
if not (_URL and _KEY):
    raise RuntimeError("SUPABASE_URL and SUPABASE_SERVICE_KEY must be set.")
db: Client = create_client(_URL, _KEY)

router = APIRouter(prefix="/api/tracker", tags=["project-tracker"])

# ─────────────────────────────────────────────────────────────────────────────
# Per-project XLSX template — strict header order
# ─────────────────────────────────────────────────────────────────────────────
MAX_TEAM = 4
TEAM_COLUMNS: list[str] = []
for i in range(1, MAX_TEAM + 1):
    TEAM_COLUMNS += [f"student_{i}_usn", f"student_{i}_name"]

TEMPLATE_COLUMNS: list[str] = [
    "project_title", "github_link", "guide_name", *TEAM_COLUMNS,
]


# ── Pydantic schemas ─────────────────────────────────────────────────────────
class BatchIn(BaseModel):
    batch_name: str
    year: int


class SemesterIn(BaseModel):
    sem_number: int


class CourseIn(BaseModel):
    course_name: str


class ProjectIn(BaseModel):
    title: Optional[str] = None
    github: Optional[str] = None
    guide: Optional[str] = None


class ProjectUpdate(BaseModel):
    title: Optional[str] = None
    github: Optional[str] = None
    guide: Optional[str] = None


class StudentUpdate(BaseModel):
    usn: Optional[str] = None
    name: Optional[str] = None


class CoordinatorIn(BaseModel):
    user_id: str   # Supabase auth user UUID


# ── Helpers ──────────────────────────────────────────────────────────────────
def _str(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _course_id_for_project(project_id: str) -> str:
    p = db.table("project").select("course_id").eq("project_id", project_id).execute()
    if not p.data:
        raise HTTPException(404, "Project not found.")
    return p.data[0]["course_id"]


# ─────────────────────────────────────────────────────────────────────────────
# Navigation — read endpoints (any authenticated user)
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/batches")
def list_batches(_user=Depends(verify_token)):
    return db.table("batch").select("*").order("year", desc=True).execute().data


@router.get("/batches/{batch_id}/semesters")
def list_semesters(batch_id: str, _user=Depends(verify_token)):
    return (
        db.table("semester").select("*")
        .eq("batch_id", batch_id).order("sem_number").execute().data
    )


@router.get("/semesters/{semester_id}/courses")
def list_courses(semester_id: str, _user=Depends(verify_token)):
    return (
        db.table("course").select("*")
        .eq("semester_id", semester_id).order("course_name").execute().data
    )


@router.get("/courses/{course_id}/projects")
def list_projects(course_id: str, _user=Depends(verify_token)):
    return (
        db.table("project")
        .select("project_id, title, github, guide, student(student_id, usn, name)")
        .eq("course_id", course_id).order("title").execute().data
    )


@router.get("/projects/{project_id}")
def get_project(project_id: str, _user=Depends(verify_token)):
    res = (
        db.table("project")
        .select(
            "project_id, title, github, guide, course_id, "
            "student(student_id, usn, name), "
            "course(course_id, course_name, semester_id, "
            "semester(semester_id, sem_number, batch_id, "
            "batch(batch_id, batch_name, year)))"
        )
        .eq("project_id", project_id).execute()
    )
    if not res.data:
        raise HTTPException(404, "Project not found.")
    return res.data[0]


# ─────────────────────────────────────────────────────────────────────────────
# Write endpoints — admin (auth-only) for tree creation,
# coordinator for project edits
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/batches")
def create_batch(body: BatchIn, _user=Depends(verify_token)):
    return db.table("batch").insert(body.dict()).execute().data[0]


@router.post("/batches/{batch_id}/semesters")
def create_semester(batch_id: str, body: SemesterIn, _user=Depends(verify_token)):
    return db.table("semester").insert(
        {"batch_id": batch_id, **body.dict()}
    ).execute().data[0]


@router.post("/semesters/{semester_id}/courses")
def create_course(semester_id: str, body: CourseIn, _user=Depends(verify_token)):
    return db.table("course").insert(
        {"semester_id": semester_id, **body.dict()}
    ).execute().data[0]


@router.post("/courses/{course_id}/coordinators")
def add_coordinator(course_id: str, body: CoordinatorIn, _user=Depends(verify_token)):
    """Assign a Supabase auth user as the course coordinator.
    Anyone authenticated can run this for now — tighten to admin role later.
    """
    return db.table("coordinator").upsert(
        {"user_id": body.user_id, "course_id": course_id}
    ).execute().data[0]


@router.get("/courses/{course_id}/coordinators")
def list_coordinators(course_id: str, _user=Depends(verify_token)):
    return db.table("coordinator").select("*").eq("course_id", course_id).execute().data


@router.post("/courses/{course_id}/projects")
def create_project(course_id: str, body: ProjectIn, user=Depends(verify_token)):
    require_coordinator_for_course(course_id, user)
    payload = {"course_id": course_id, **{k: v for k, v in body.dict().items() if v is not None}}
    return db.table("project").insert(payload).execute().data[0]


@router.put("/projects/{project_id}")
def update_project(project_id: str, body: ProjectUpdate, user=Depends(verify_token)):
    require_coordinator_for_project(project_id, user)
    patch = {k: v for k, v in body.dict().items() if v is not None}
    if not patch:
        raise HTTPException(400, "No fields provided.")
    res = db.table("project").update(patch).eq("project_id", project_id).execute()
    if not res.data:
        raise HTTPException(404, "Project not found.")
    return res.data[0]


@router.delete("/projects/{project_id}")
def delete_project(project_id: str, user=Depends(verify_token)):
    require_coordinator_for_project(project_id, user)
    res = db.table("project").delete().eq("project_id", project_id).execute()
    if not res.data:
        raise HTTPException(404, "Project not found.")
    return {"message": "Project deleted."}


@router.put("/students/{student_id}")
def update_student(student_id: str, body: StudentUpdate, user=Depends(verify_token)):
    # Resolve project → course → coordinator check
    s = db.table("student").select("project_id").eq("student_id", student_id).execute()
    if not s.data:
        raise HTTPException(404, "Student not found.")
    require_coordinator_for_project(s.data[0]["project_id"], user)

    patch = {k: v for k, v in body.dict().items() if v is not None}
    if not patch:
        raise HTTPException(400, "No fields provided.")
    res = db.table("student").update(patch).eq("student_id", student_id).execute()
    return res.data[0]


@router.delete("/students/{student_id}")
def delete_student(student_id: str, user=Depends(verify_token)):
    s = db.table("student").select("project_id").eq("student_id", student_id).execute()
    if not s.data:
        raise HTTPException(404, "Student not found.")
    require_coordinator_for_project(s.data[0]["project_id"], user)
    db.table("student").delete().eq("student_id", student_id).execute()
    return {"message": "Student deleted."}


# ─────────────────────────────────────────────────────────────────────────────
# Per-project XLSX upload — strict header validation, replaces existing team
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/projects/{project_id}/upload-excel")
async def upload_project_excel(
    project_id: str,
    file: UploadFile = File(...),
    user=Depends(verify_token),
):
    require_coordinator_for_project(project_id, user)

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx files are accepted.")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"Could not parse the file ({exc}).")

    ws = wb.active

    # ── Validate header — exact column order ────────────────────────────────
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    if len(header) != len(TEMPLATE_COLUMNS):
        raise HTTPException(
            400,
            f"Header has {len(header)} columns; expected {len(TEMPLATE_COLUMNS)}. "
            f"Format changed — re-download the template.",
        )
    for idx, (got, want) in enumerate(zip(header, TEMPLATE_COLUMNS), start=1):
        if got != want:
            raise HTTPException(
                400,
                f"Column {idx} must be '{want}' but got '{got}'. "
                f"Format changed — re-download the template.",
            )

    # ── Pull the first non-empty data row ────────────────────────────────────
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    row = next(
        (r for r in data_rows if any(v is not None and str(v).strip() != "" for v in r)),
        None,
    )
    if row is None:
        raise HTTPException(400, "Sheet has no data rows.")

    cells = dict(zip(TEMPLATE_COLUMNS, row))

    # ── Update project metadata (only overwrite when a value was provided) ──
    proj_patch = {
        k: _str(cells.get(src))
        for k, src in [("title", "project_title"), ("github", "github_link"), ("guide", "guide_name")]
        if _str(cells.get(src)) is not None
    }
    if proj_patch:
        db.table("project").update(proj_patch).eq("project_id", project_id).execute()

    # ── Replace team — both USN and name required, otherwise slot ignored ───
    db.table("student").delete().eq("project_id", project_id).execute()

    new_students: list[dict] = []
    skipped_slots: list[int] = []
    for i in range(1, MAX_TEAM + 1):
        usn = _str(cells.get(f"student_{i}_usn"))
        name = _str(cells.get(f"student_{i}_name"))
        if usn and name:
            new_students.append({"project_id": project_id, "usn": usn, "name": name})
        elif usn or name:
            skipped_slots.append(i)  # partial → skip silently but report

    inserted = 0
    if new_students:
        ins = db.table("student").insert(new_students).execute()
        inserted = len(ins.data or [])

    return {
        "project": db.table("project").select("*").eq("project_id", project_id).execute().data[0],
        "students_inserted": inserted,
        "partial_slots_skipped": skipped_slots,
        "applied_project_fields": list(proj_patch.keys()),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Download Template — per-project layout
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/download-template")
def download_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    hdr_fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
    hdr_font = Font(color="F5A623", bold=True)

    ws.append(TEMPLATE_COLUMNS)
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    # One example row showing a 3-person team (4th slot blank, still accepted)
    ws.append([
        "Image Classifier",
        "https://github.com/team/ml-proj",
        "Dr. Smith",
        "1CS21CS001", "Alice Kumar",
        "1CS21CS002", "Bob Nair",
        "1CS21CS003", "Carol Menon",
        "", "",
    ])

    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(width + 4, 16)

    iws = wb.create_sheet("Instructions")
    iws.append(["Column", "Required?", "Notes"])
    for cell in iws[1]:
        cell.font = Font(bold=True)
    for row in [
        ("project_title", "Optional", "Filled into the project record."),
        ("github_link",   "Optional", "Leave empty if repo isn't created yet."),
        ("guide_name",    "Optional", "Faculty guide / mentor."),
        ("student_N_usn", "Optional", "Slot is used only if BOTH usn and name are filled."),
        ("student_N_name","Optional", "Up to 4 students per team."),
        ("",              "",         "Re-download this template before uploading — format is validated."),
    ]:
        iws.append(row)
    for col in iws.columns:
        width = max(len(str(c.value or "")) for c in col)
        iws.column_dimensions[col[0].column_letter].width = max(width + 4, 18)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=project_template.xlsx"},
    )


# ─────────────────────────────────────────────────────────────────────────────
# README passthrough — used by ProjectDetailScreen
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/projects/{project_id}/readme")
def get_project_readme(project_id: str, _user=Depends(verify_token)):
    """Returns { content, found }. Never throws on a missing/private repo."""
    from github_utils import parse_repo_url, fetch_readme

    proj = db.table("project").select("github").eq("project_id", project_id).execute()
    if not proj.data:
        raise HTTPException(404, "Project not found.")

    parsed = parse_repo_url(proj.data[0].get("github"))
    if not parsed:
        return {"found": False, "content": "", "reason": "no-github-link"}

    content = fetch_readme(*parsed)
    if not content:
        return {"found": False, "content": "", "reason": "private-or-missing"}
    return {"found": True, "content": content}
